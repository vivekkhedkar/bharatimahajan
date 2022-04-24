import requests
# from bs4 import BeautifulSoup
import os
# import pandas as pd
from openpyxl import Workbook, load_workbook

# workbook = Workbook()
# sheet = workbook.active



os.chdir("C:/cs50w/Bharati/duda_images")
wb = load_workbook('images.xlsx')
sheet = wb['Sheet1']

# link2='https://irp.cdn-website.com/e415ee9d/dms3rep/multi/yy.jpg'
# with open('yyyyyy.jpg','wb') as f:
#     im = requests.get(link2)
#     f.write(im.content)


for cell in sheet:
    imageName = cell[0].value
    link = 'https://irp.cdn-website.com/e415ee9d/dms3rep/multi/'+ imageName
    with open (imageName,'wb') as f: # using 'with open' you dont have to close the file manually by writing f.close()
        im = requests.get(link)
        f.write(im.content)



    







